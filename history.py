import json
import csv
import os
from datetime import datetime
from typing import Optional
from models import PublishRecord, PublishStatus
from compliance_log import ComplianceLogger


class HistoryManager:
    def __init__(self, config: dict, logger: ComplianceLogger):
        self.data_dir = config.get("system", {}).get("data_dir", "./data")
        self.export_dir = config.get("system", {}).get("export_dir", "./exports")
        self.batch_size = config.get("history", {}).get("batch_size", 1000)
        self.channel_names = {"app": "APP", "phone": "电话", "wechat": "微信", "mini_program": "小程序"}
        self.status_names = {
            "pending_check": "待检查",
            "check_failed": "检查失败",
            "pending_approval": "待审批",
            "approval_rejected": "审批拒绝",
            "gray_releasing": "灰度发布中",
            "published": "已发布",
            "rolled_back": "已回滚",
        }
        self.risk_names = {
            "routine": "常规话术更新",
            "emergency_compliance": "紧急合规整改",
            "complaint_outbreak": "客户投诉爆发",
        }
        os.makedirs(self.data_dir, exist_ok=True)
        os.makedirs(self.export_dir, exist_ok=True)
        self.logger = logger

    def save_record(self, record: PublishRecord):
        records_file = os.path.join(self.data_dir, "publish_records.jsonl")
        data = self._record_to_dict(record)
        with open(records_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(data, ensure_ascii=False) + "\n")

    def query_records(
        self,
        publish_time_start: Optional[datetime] = None,
        publish_time_end: Optional[datetime] = None,
        channel: Optional[str] = None,
        business_type: Optional[str] = None,
        version: Optional[str] = None,
        status: Optional[str] = None,
        risk_level: Optional[str] = None,
        operator: Optional[str] = None,
        limit: int = 100,
        offset: int = 0,
    ) -> list:
        records = self._load_all_records()
        filtered = self._filter_records(
            records,
            publish_time_start=publish_time_start,
            publish_time_end=publish_time_end,
            channel=channel,
            business_type=business_type,
            version=version,
            status=status,
            risk_level=risk_level,
            operator=operator,
        )
        return filtered[offset : offset + limit]

    def count_records(
        self,
        publish_time_start: Optional[datetime] = None,
        publish_time_end: Optional[datetime] = None,
        channel: Optional[str] = None,
        business_type: Optional[str] = None,
        version: Optional[str] = None,
        status: Optional[str] = None,
        risk_level: Optional[str] = None,
        operator: Optional[str] = None,
    ) -> int:
        records = self._load_all_records()
        filtered = self._filter_records(
            records,
            publish_time_start=publish_time_start,
            publish_time_end=publish_time_end,
            channel=channel,
            business_type=business_type,
            version=version,
            status=status,
            risk_level=risk_level,
            operator=operator,
        )
        return len(filtered)

    def batch_export(
        self,
        records: list,
        format: str = "csv",
        filename: Optional[str] = None,
        operator: str = "system",
    ) -> str:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "xlsx" if format == "excel" else format
            filename = f"export_{timestamp}.{ext}"

        filepath = os.path.join(self.export_dir, filename)

        if format == "excel":
            filepath = self._export_excel(records, filepath)
        else:
            filepath = self._export_csv(records, filepath)

        self.logger.log(
            "batch_export",
            operator,
            "history",
            f"批量导出{len(records)}条记录: {filepath}",
            {"format": format, "count": len(records)},
        )
        return filepath

    def _load_all_records(self) -> list:
        records_file = os.path.join(self.data_dir, "publish_records.jsonl")
        if not os.path.exists(records_file):
            return []
        records = []
        with open(records_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
        return records

    def _filter_records(
        self,
        records: list,
        publish_time_start=None,
        publish_time_end=None,
        channel=None,
        business_type=None,
        version=None,
        status=None,
        risk_level=None,
        operator=None,
    ) -> list:
        filtered = []
        for r in records:
            if publish_time_start:
                ts = datetime.fromisoformat(r.get("submitted_at", ""))
                if ts < publish_time_start:
                    continue
            if publish_time_end:
                ts = datetime.fromisoformat(r.get("submitted_at", ""))
                if ts > publish_time_end:
                    continue
            if channel and r.get("script_version", {}).get("channel") != channel:
                continue
            if business_type and r.get("script_version", {}).get("business_type") != business_type:
                continue
            if version and r.get("script_version", {}).get("version") != version:
                continue
            if status and r.get("status") != status:
                continue
            if risk_level and r.get("risk_level") != risk_level:
                continue
            if operator and r.get("operator") != operator:
                continue
            filtered.append(r)
        return filtered

    def _export_csv(self, records: list, filepath: str) -> str:
        headers = [
            "发布ID",
            "话术版本",
            "业务类型",
            "服务渠道",
            "风险级别",
            "发布状态",
            "提交人",
            "提交时间",
            "发布时间",
            "回滚时间",
        ]
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for r in records:
                sv = r.get("script_version", {})
                writer.writerow(
                    [
                        r.get("publish_id", ""),
                        sv.get("version", ""),
                        sv.get("business_type", ""),
                        self.channel_names.get(sv.get("channel", ""), sv.get("channel", "")),
                        self.risk_names.get(r.get("risk_level", ""), r.get("risk_level", "")),
                        self.status_names.get(r.get("status", ""), r.get("status", "")),
                        r.get("operator", ""),
                        r.get("submitted_at", ""),
                        r.get("published_at", ""),
                        r.get("rolled_back_at", ""),
                    ]
                )
        return filepath

    def _export_excel(self, records: list, filepath: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side
        except ImportError:
            return self._export_csv(records, filepath.replace(".xlsx", ".csv"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发布记录"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "发布ID",
            "话术版本",
            "业务类型",
            "服务渠道",
            "风险级别",
            "发布状态",
            "提交人",
            "提交时间",
            "发布时间",
            "回滚时间",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, r in enumerate(records, 2):
            sv = r.get("script_version", {})
            row_data = [
                r.get("publish_id", ""),
                sv.get("version", ""),
                sv.get("business_type", ""),
                self.channel_names.get(sv.get("channel", ""), sv.get("channel", "")),
                self.risk_names.get(r.get("risk_level", ""), r.get("risk_level", "")),
                self.status_names.get(r.get("status", ""), r.get("status", "")),
                r.get("operator", ""),
                r.get("submitted_at", ""),
                r.get("published_at", ""),
                r.get("rolled_back_at", ""),
            ]
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=val).border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        wb.save(filepath)
        return filepath

    def _record_to_dict(self, record: PublishRecord) -> dict:
        rollback = None
        if record.rollback_report:
            rollback = {
                "rollback_id": record.rollback_report.rollback_id,
                "trigger": record.rollback_report.trigger.value,
                "restored_version": record.rollback_report.restored_version,
                "channel_impact": record.rollback_report.channel_impact,
                "violation_reasons": record.rollback_report.violation_reasons,
                "complaint_stats": record.rollback_report.complaint_stats,
                "rolled_back_at": record.rollback_report.rolled_back_at.isoformat(),
                "notified_roles": record.rollback_report.notified_roles,
            }
        return {
            "publish_id": record.publish_id,
            "script_version": {
                "version": record.script_version.version,
                "content": record.script_version.content[:200],
                "business_type": record.script_version.business_type,
                "channel": record.script_version.channel,
                "created_at": record.script_version.created_at.isoformat(),
                "checksum": record.script_version.checksum,
            },
            "risk_level": record.risk_level.value,
            "status": record.status.value,
            "operator": record.operator,
            "submitted_at": record.submitted_at.isoformat(),
            "published_at": record.published_at.isoformat() if record.published_at else None,
            "rolled_back_at": record.rolled_back_at.isoformat() if record.rolled_back_at else None,
            "rollback_report": rollback,
        }
