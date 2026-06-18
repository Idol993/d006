import json
import os
import csv
from datetime import datetime, timedelta
from typing import Optional
from collections import defaultdict
from models import PublishRecord, PublishStatus, WeeklyStats, RiskLevel, Channel
from compliance_log import ComplianceLogger


class ReportManager:
    def __init__(self, config: dict, logger: ComplianceLogger):
        self.config = config.get("report", {})
        self.report_dir = config.get("system", {}).get("report_dir", "./reports")
        self.data_dir = config.get("system", {}).get("data_dir", "./data")
        os.makedirs(self.report_dir, exist_ok=True)
        os.makedirs(self.data_dir, exist_ok=True)
        self.logger = logger
        self.channel_names = {"app": "APP", "phone": "电话", "wechat": "微信", "mini_program": "小程序"}
        self.risk_names = {
            "routine": "常规话术更新",
            "emergency_compliance": "紧急合规整改",
            "complaint_outbreak": "客户投诉爆发",
        }

    def generate_weekly_stats(
        self, records: list, period_start: Optional[datetime] = None, operator: str = "system"
    ) -> WeeklyStats:
        if period_start is None:
            today = datetime.now()
            period_start = today - timedelta(days=today.weekday())
        period_end = period_start + timedelta(days=6)

        stats = WeeklyStats(period_start=period_start, period_end=period_end)

        period_records = [
            r
            for r in records
            if period_start <= r.submitted_at <= period_end
        ]

        stats.total_publishes = len(period_records)
        stats.success_publishes = sum(
            1 for r in period_records if r.status == PublishStatus.PUBLISHED
        )
        stats.publish_success_rate = (
            stats.success_publishes / stats.total_publishes
            if stats.total_publishes > 0
            else 0.0
        )
        stats.rollback_count = sum(
            1 for r in period_records if r.status == PublishStatus.ROLLED_BACK
        )
        satisfaction_values = []
        for r in period_records:
            if r.monitor_snapshots:
                avg_resp = sum(s.response_rate for s in r.monitor_snapshots) / len(
                    r.monitor_snapshots
                )
                satisfaction_values.append(avg_resp)
        stats.avg_customer_satisfaction = (
            sum(satisfaction_values) / len(satisfaction_values)
            if satisfaction_values
            else 0.0
        )

        by_channel = defaultdict(lambda: {"total": 0, "success": 0, "rollback": 0})
        for r in period_records:
            ch = r.script_version.channel
            by_channel[ch]["total"] += 1
            if r.status == PublishStatus.PUBLISHED:
                by_channel[ch]["success"] += 1
            elif r.status == PublishStatus.ROLLED_BACK:
                by_channel[ch]["rollback"] += 1
        stats.by_channel = dict(by_channel)

        by_risk = defaultdict(lambda: {"total": 0, "success": 0, "rollback": 0})
        for r in period_records:
            rk = r.risk_level.value
            by_risk[rk]["total"] += 1
            if r.status == PublishStatus.PUBLISHED:
                by_risk[rk]["success"] += 1
            elif r.status == PublishStatus.ROLLED_BACK:
                by_risk[rk]["rollback"] += 1
        stats.by_risk_level = dict(by_risk)

        by_biz = defaultdict(lambda: {"total": 0, "success": 0, "rollback": 0})
        for r in period_records:
            bt = r.script_version.business_type
            by_biz[bt]["total"] += 1
            if r.status == PublishStatus.PUBLISHED:
                by_biz[bt]["success"] += 1
            elif r.status == PublishStatus.ROLLED_BACK:
                by_biz[bt]["rollback"] += 1
        stats.by_business_type = dict(by_biz)

        self.logger.log(
            "weekly_stats_generated",
            operator,
            "weekly_report",
            f"周报统计生成: {period_start.strftime('%Y-%m-%d')} ~ {period_end.strftime('%Y-%m-%d')}",
            {
                "total": stats.total_publishes,
                "success": stats.success_publishes,
                "rollback": stats.rollback_count,
                "success_rate": f"{stats.publish_success_rate:.1%}",
            },
        )
        return stats

    def generate_pdf_report(self, stats: WeeklyStats, operator: str = "system") -> str:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
                Image,
            )
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            self.logger.log(
                "pdf_report_error",
                operator,
                "weekly_report",
                "reportlab未安装，使用文本格式替代",
            )
            return self._generate_text_report(stats, operator)

        chart_path = self._generate_charts(stats)
        filename = f"weekly_report_{stats.period_start.strftime('%Y%m%d')}.pdf"
        filepath = os.path.join(self.report_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(
            Paragraph(
                f"金融智能客服合规话术管理 - 周报<br/>"
                f"统计周期: {stats.period_start.strftime('%Y-%m-%d')} ~ {stats.period_end.strftime('%Y-%m-%d')}",
                styles["Title"],
            )
        )
        elements.append(Spacer(1, 0.5 * cm))

        summary_data = [
            ["指标", "数值"],
            ["话术发布总数", str(stats.total_publishes)],
            ["发布成功数", str(stats.success_publishes)],
            ["发布成功率", f"{stats.publish_success_rate:.1%}"],
            ["回滚次数", str(stats.rollback_count)],
            ["平均客户满意度", f"{stats.avg_customer_satisfaction:.1%}"],
        ]
        summary_table = Table(summary_data)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 1 * cm))

        if stats.by_channel:
            ch_data = [["渠道", "总数", "成功", "回滚", "成功率"]]
            for ch, ch_stats in stats.by_channel.items():
                rate = ch_stats["success"] / ch_stats["total"] if ch_stats["total"] > 0 else 0
                ch_data.append(
                    [
                        self.channel_names.get(ch, ch),
                        str(ch_stats["total"]),
                        str(ch_stats["success"]),
                        str(ch_stats["rollback"]),
                        f"{rate:.1%}",
                    ]
                )
            ch_table = Table(ch_data)
            ch_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            elements.append(ch_table)
            elements.append(Spacer(1, 1 * cm))

        if chart_path and os.path.exists(chart_path):
            elements.append(Paragraph("服务趋势图表", styles["Heading2"]))
            elements.append(Image(chart_path, width=15 * cm, height=10 * cm))

        doc.build(elements)

        self.logger.log(
            "pdf_report_generated",
            operator,
            "weekly_report",
            f"PDF周报已生成: {filepath}",
            {"filename": filename},
        )
        return filepath

    def generate_excel_report(self, stats: WeeklyStats, operator: str = "system") -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.logger.log(
                "excel_report_error",
                operator,
                "weekly_report",
                "openpyxl未安装，使用CSV格式替代",
            )
            return self._generate_csv_report(stats, operator)

        filename = f"weekly_operation_{stats.period_start.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.report_dir, filename)

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws_summary = wb.active
        ws_summary.title = "发布概览"
        headers = ["指标", "数值"]
        for col, h in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        summary_rows = [
            ("统计周期", f"{stats.period_start.strftime('%Y-%m-%d')} ~ {stats.period_end.strftime('%Y-%m-%d')}"),
            ("话术发布总数", stats.total_publishes),
            ("发布成功数", stats.success_publishes),
            ("发布成功率", f"{stats.publish_success_rate:.1%}"),
            ("回滚次数", stats.rollback_count),
            ("平均客户满意度", f"{stats.avg_customer_satisfaction:.1%}"),
        ]
        for row_idx, (label, value) in enumerate(summary_rows, 2):
            ws_summary.cell(row=row_idx, column=1, value=label).border = thin_border
            ws_summary.cell(row=row_idx, column=2, value=value).border = thin_border

        ws_channel = wb.create_sheet("渠道分析")
        ch_headers = ["渠道", "发布总数", "成功数", "回滚数", "成功率"]
        for col, h in enumerate(ch_headers, 1):
            cell = ws_channel.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        row_idx = 2
        for ch, ch_stats in stats.by_channel.items():
            rate = ch_stats["success"] / ch_stats["total"] if ch_stats["total"] > 0 else 0
            ws_channel.cell(row=row_idx, column=1, value=self.channel_names.get(ch, ch)).border = thin_border
            ws_channel.cell(row=row_idx, column=2, value=ch_stats["total"]).border = thin_border
            ws_channel.cell(row=row_idx, column=3, value=ch_stats["success"]).border = thin_border
            ws_channel.cell(row=row_idx, column=4, value=ch_stats["rollback"]).border = thin_border
            ws_channel.cell(row=row_idx, column=5, value=f"{rate:.1%}").border = thin_border
            row_idx += 1

        ws_risk = wb.create_sheet("风险级别分析")
        risk_headers = ["风险级别", "发布总数", "成功数", "回滚数"]
        for col, h in enumerate(risk_headers, 1):
            cell = ws_risk.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        row_idx = 2
        for rk, rk_stats in stats.by_risk_level.items():
            ws_risk.cell(row=row_idx, column=1, value=self.risk_names.get(rk, rk)).border = thin_border
            ws_risk.cell(row=row_idx, column=2, value=rk_stats["total"]).border = thin_border
            ws_risk.cell(row=row_idx, column=3, value=rk_stats["success"]).border = thin_border
            ws_risk.cell(row=row_idx, column=4, value=rk_stats["rollback"]).border = thin_border
            row_idx += 1

        for ws in [ws_summary, ws_channel, ws_risk]:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        wb.save(filepath)

        self.logger.log(
            "excel_report_generated",
            operator,
            "weekly_report",
            f"Excel周报已生成: {filepath}",
            {"filename": filename},
        )
        return filepath

    def _generate_charts(self, stats: WeeklyStats) -> Optional[str]:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
            plt.rcParams["axes.unicode_minus"] = False
        except ImportError:
            return None

        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle("金融智能客服合规话术 - 服务趋势图表", fontsize=16)

        if stats.by_channel:
            channels = [self.channel_names.get(ch, ch) for ch in stats.by_channel.keys()]
            success_counts = [v["success"] for v in stats.by_channel.values()]
            rollback_counts = [v["rollback"] for v in stats.by_channel.values()]
            x = range(len(channels))
            axes[0, 0].bar(x, success_counts, label="成功", color="#4472C4")
            axes[0, 0].bar(x, rollback_counts, bottom=success_counts, label="回滚", color="#ED7D31")
            axes[0, 0].set_xticks(x)
            axes[0, 0].set_xticklabels(channels)
            axes[0, 0].set_title("各渠道发布统计")
            axes[0, 0].legend()

        if stats.by_risk_level:
            risk_labels = [self.risk_names.get(rk, rk) for rk in stats.by_risk_level.keys()]
            risk_counts = [v["total"] for v in stats.by_risk_level.values()]
            axes[0, 1].pie(risk_counts, labels=risk_labels, autopct="%1.1f%%", startangle=90)
            axes[0, 1].set_title("风险级别分布")

        axes[1, 0].bar(
            ["发布成功率", "客户满意度"],
            [stats.publish_success_rate, stats.avg_customer_satisfaction],
            color=["#4472C4", "#70AD47"],
        )
        axes[1, 0].set_ylim(0, 1.1)
        axes[1, 0].set_title("关键指标")
        axes[1, 0].axhline(y=0.85, color="r", linestyle="--", label="阈值")
        axes[1, 0].legend()

        axes[1, 1].text(
            0.5,
            0.5,
            f"回滚次数: {stats.rollback_count}\n\n"
            f"发布总数: {stats.total_publishes}\n\n"
            f"成功数: {stats.success_publishes}",
            ha="center",
            va="center",
            fontsize=14,
            transform=axes[1, 1].transAxes,
        )
        axes[1, 1].set_title("本周概览")

        plt.tight_layout()
        chart_path = os.path.join(
            self.report_dir,
            f"chart_{stats.period_start.strftime('%Y%m%d')}.png",
        )
        fig.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        return chart_path

    def _generate_text_report(self, stats: WeeklyStats, operator: str) -> str:
        filename = f"weekly_report_{stats.period_start.strftime('%Y%m%d')}.txt"
        filepath = os.path.join(self.report_dir, filename)
        lines = [
            "=" * 60,
            "金融智能客服合规话术管理 - 周报",
            f"统计周期: {stats.period_start.strftime('%Y-%m-%d')} ~ {stats.period_end.strftime('%Y-%m-%d')}",
            "=" * 60,
            "",
            f"话术发布总数: {stats.total_publishes}",
            f"发布成功数:   {stats.success_publishes}",
            f"发布成功率:   {stats.publish_success_rate:.1%}",
            f"回滚次数:     {stats.rollback_count}",
            f"客户满意度:   {stats.avg_customer_satisfaction:.1%}",
            "",
            "--- 渠道分析 ---",
        ]
        for ch, ch_stats in stats.by_channel.items():
            rate = ch_stats["success"] / ch_stats["total"] if ch_stats["total"] > 0 else 0
            lines.append(
                f"  {self.channel_names.get(ch, ch)}: "
                f"总数{ch_stats['total']}, 成功{ch_stats['success']}, "
                f"回滚{ch_stats['rollback']}, 成功率{rate:.1%}"
            )
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        return filepath

    def _generate_csv_report(self, stats: WeeklyStats, operator: str) -> str:
        filename = f"weekly_operation_{stats.period_start.strftime('%Y%m%d')}.csv"
        filepath = os.path.join(self.report_dir, filename)
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["指标", "数值"])
            writer.writerow(["话术发布总数", stats.total_publishes])
            writer.writerow(["发布成功数", stats.success_publishes])
            writer.writerow(["发布成功率", f"{stats.publish_success_rate:.1%}"])
            writer.writerow(["回滚次数", stats.rollback_count])
            writer.writerow(["平均客户满意度", f"{stats.avg_customer_satisfaction:.1%}"])
        return filepath
